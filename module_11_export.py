"""
Module 11: CSV and Excel Export (v2 — Fully Automated)

Writes the final output file.
Column order matches the CRM schema exactly (21 columns).
Naming convention: HotelName_Year.csv / HotelName_Year.xlsx
"""

import csv
import os
import re
from datetime import datetime
from typing import List, Optional

from module_09_mapping import CSV_COLUMNS, CSVRow


def export_csv(
    rows: List[CSVRow],
    hotel_name: str,
    contract_year: Optional[str] = None,
    output_dir: str = ".",
    filename_override: Optional[str] = None,
) -> str:
    """Write rows to CSV. Returns the full path of the created file."""
    os.makedirs(output_dir, exist_ok=True)
    filepath = _build_filepath(hotel_name, contract_year, output_dir,
                               filename_override, ext="csv")

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_dict())

    print(f"✅ CSV exported: {filepath}  ({len(rows)} rows)")
    return filepath


def export_excel(
    rows: List[CSVRow],
    hotel_name: str,
    contract_year: Optional[str] = None,
    output_dir: str = ".",
    filename_override: Optional[str] = None,
) -> str:
    """Write rows to Excel (.xlsx). Returns the full path of the created file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

    os.makedirs(output_dir, exist_ok=True)
    filepath = _build_filepath(hotel_name, contract_year, output_dir,
                               filename_override, ext="xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Rates"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", wrap_text=True)

    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        d = row.to_dict()
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=d.get(col_name, ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"✅ Excel exported: {filepath}  ({len(rows)} rows)")
    return filepath


def export_both(
    rows: List[CSVRow],
    hotel_name: str,
    contract_year: Optional[str] = None,
    output_dir: str = ".",
) -> dict:
    """Export both CSV and Excel. Returns dict with 'csv' and 'xlsx' paths."""
    return {
        "csv":  export_csv(rows, hotel_name, contract_year, output_dir),
        "xlsx": export_excel(rows, hotel_name, contract_year, output_dir),
    }


def preview_csv(rows: List[CSVRow], max_rows: int = 8) -> None:
    """Print a preview table to stdout."""
    cols_preview = ["room_name", "min_age", "max_age", "start_date", "end_date",
                    "cost", "single_supplement", "hb_supplement", "fb_supplement"]
    header = " | ".join(f"{c:<18}" for c in cols_preview)
    sep    = "─" * len(header)
    print(f"\n{sep}\n{header}\n{sep}")
    for row in rows[:max_rows]:
        d = row.to_dict()
        print(" | ".join(f"{str(d.get(c,'')):<18}" for c in cols_preview))
    if len(rows) > max_rows:
        print(f"  … {len(rows) - max_rows} more rows not shown")
    print(f"{sep}\n")


def _build_filepath(hotel_name, contract_year, output_dir, override, ext):
    if override:
        fname = override if override.endswith(f".{ext}") else f"{override}.{ext}"
    else:
        safe = _safe_filename(hotel_name)
        year = contract_year or datetime.now().strftime("%Y")
        fname = f"{safe}_{year}.{ext}"
    return os.path.join(output_dir, fname)


def _safe_filename(name: str) -> str:
    safe = re.sub(r"[^\w\s\-]", "", name)
    safe = re.sub(r"\s+", "_", safe.strip())
    return safe[:60] or "Hotel"
